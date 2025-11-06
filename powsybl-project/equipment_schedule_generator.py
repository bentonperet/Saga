"""
Equipment Schedule Generator for Electrical Single-Line Diagrams
Creates professional Excel spreadsheets with equipment specifications, ratings, and BOM
"""

import json
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_equipment_schedule(metadata_file, output_file=None, project_name="SAGA PRYOR DATA CENTER"):
    """
    Generate comprehensive equipment schedule from metadata
    
    Args:
        metadata_file: Path to JSON metadata file
        output_file: Path to output Excel file (optional)
        project_name: Project name for header
    
    Returns:
        Path to created Excel file
    """
    # Load metadata
    with open(metadata_file, 'r') as f:
        data = json.load(f)
    
    if output_file is None:
        output_file = Path(metadata_file).with_suffix('.xlsx')
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subheader_font = Font(color="FFFFFF", bold=True, size=10)
    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === SHEET 1: SUMMARY ===
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary['A1'] = project_name
    ws_summary['A1'].font = title_font
    ws_summary['A2'] = "ELECTRICAL EQUIPMENT SCHEDULE"
    ws_summary['A2'].font = Font(bold=True, size=12)
    ws_summary['A3'] = f"Generated: {datetime.now().strftime('%m/%d/%Y %H:%M')}"
    
    # Summary table
    row = 5
    ws_summary[f'A{row}'] = "SYSTEM SUMMARY"
    ws_summary[f'A{row}'].font = bold_font
    row += 1
    
    gen_rating = f"{data['generators'][0].get('kw', 'N/A')}kW" if data.get('generators') else "N/A"
    tx_rating = f"{data['transformers'][0].get('rated_s', 'N/A')}kVA" if data.get('transformers') else "N/A"
    ups_rating = f"{data['ups'][0].get('kw', 'N/A')}kW" if data.get('ups') else "N/A"
    
    summary_data = [
        ("Utility Feeds", len(data.get('utility_feeds', [])), "13.8 kV"),
        ("Generators", len(data.get('generators', [])), f"{gen_rating} each"),
        ("MV Switchboards", len(data.get('mv_switchboards', [])), "13.8 kV"),
        ("Ring Main Units", len(data.get('rmu', [])), "630A"),
        ("Transformers", len(data.get('transformers', [])), tx_rating),
        ("LV Switchboards", len(data.get('lv_swbd', [])), "480V"),
        ("UPS Systems", len(data.get('ups', [])), ups_rating),
    ]
    
    ws_summary[f'A{row}'] = "Category"
    ws_summary[f'B{row}'] = "Quantity"
    ws_summary[f'C{row}'] = "Rating"
    for cell in [f'A{row}', f'B{row}', f'C{row}']:
        ws_summary[cell].font = bold_font
        ws_summary[cell].fill = subheader_fill
        ws_summary[cell].font = subheader_font
    row += 1
    
    for category, qty, rating in summary_data:
        ws_summary[f'A{row}'] = category
        ws_summary[f'B{row}'] = qty
        ws_summary[f'C{row}'] = rating
        row += 1
    
    # Format columns
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 12
    ws_summary.column_dimensions['C'].width = 20
    
    # === SHEET 2: GENERATORS ===
    ws_gen = wb.create_sheet("Generators")
    
    headers = ["Tag", "Type", "Rating", "Voltage", "Frequency", "Power Factor", "Fuel", "Location"]
    row = 1
    for col, header in enumerate(headers, start=1):
        cell = ws_gen.cell(row, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for gen in data.get('generators', []):
        row += 1
        ws_gen.cell(row, 1, gen['id'])
        ws_gen.cell(row, 2, gen.get('type', 'RECIP'))
        ws_gen.cell(row, 3, f"{gen.get('kw', 'N/A')}kW / {gen.get('kva', 'N/A')}kVA")
        ws_gen.cell(row, 4, gen.get('voltage', '13.8 kV'))
        ws_gen.cell(row, 5, '60 Hz')
        ws_gen.cell(row, 6, '0.8')
        ws_gen.cell(row, 7, gen.get('fuel', 'NAT_GAS'))
        ws_gen.cell(row, 8, 'Generator Room')
    
    for col in range(1, len(headers) + 1):
        ws_gen.column_dimensions[get_column_letter(col)].width = 15
    
    # === SHEET 3: MV SWITCHBOARDS (optional) ===
    if data.get('mv_switchboards'):
        ws_mv = wb.create_sheet("MV Switchboards")
        headers = ["Tag", "Type", "Voltage", "Bus Rating", "Short Circuit", "Breaker Type", "Number of Feeders"]
        row = 1
        for col, header in enumerate(headers, start=1):
            cell = ws_mv.cell(row, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for swbd in data.get('mv_switchboards', []):
            row += 1
            ws_mv.cell(row, 1, swbd['id'])
            ws_mv.cell(row, 2, swbd.get('type', 'Metal-Clad Switchgear'))
            ws_mv.cell(row, 3, swbd.get('voltage', '13.8 kV'))
            ws_mv.cell(row, 4, swbd.get('bus_rating', '4000A'))
            ws_mv.cell(row, 5, swbd.get('short_circuit_rating', '65kAIC'))
            ws_mv.cell(row, 6, swbd.get('breaker_type', 'Vacuum'))
            ws_mv.cell(row, 7, 0)
        
        for col in range(1, len(headers) + 1):
            ws_mv.column_dimensions[get_column_letter(col)].width = 18
    
    # === SHEET 4: RMUs (optional) ===
    if data.get('rmu'):
        ws_rmu = wb.create_sheet("Ring Main Units")
        headers = ["Tag", "Type", "Rated Current", "Rated Voltage", "Short Circuit", "Switch Type", "Feeds From", "Feeds To"]
        row = 1
        for col, header in enumerate(headers, start=1):
            cell = ws_rmu.cell(row, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for i, rmu in enumerate(data.get('rmu', []), start=1):
            row += 1
            if isinstance(rmu, dict):
                ws_rmu.cell(row, 1, rmu.get('id', f'RMU_{i}'))
                ws_rmu.cell(row, 2, rmu.get('type', 'SF6 Gas Insulated'))
                ws_rmu.cell(row, 3, rmu.get('rating', '630A'))
                ws_rmu.cell(row, 4, rmu.get('voltage', '13.8 kV'))
                ws_rmu.cell(row, 5, rmu.get('short_circuit_rating', '20kA'))
                ws_rmu.cell(row, 6, rmu.get('switch_type', 'Load Break Switch'))
                ws_rmu.cell(row, 7, "MV-MSB A, MV-MSB B")
                ws_rmu.cell(row, 8, f"TX-{i}")
        
        for col in range(1, len(headers) + 1):
            ws_rmu.column_dimensions[get_column_letter(col)].width = 18
    
    # === SHEET 5: TRANSFORMERS ===
    ws_tx = wb.create_sheet("Transformers")
    
    headers = ["Tag", "Type", "Rating", "Primary Voltage", "Secondary Voltage", "Connection", "Impedance", "Cooling"]
    row = 1
    for col, header in enumerate(headers, start=1):
        cell = ws_tx.cell(row, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for tx in data.get('transformers', []):
        row += 1
        ws_tx.cell(row, 1, tx['id'])
        ws_tx.cell(row, 2, 'Dry Type')
        ws_tx.cell(row, 3, f"{tx.get('rated_s', 'N/A')}kVA")
        ws_tx.cell(row, 4, f"{tx.get('hv_voltage', 'N/A')}kV")
        ws_tx.cell(row, 5, f"{tx.get('lv_voltage', 'N/A')}kV")
        ws_tx.cell(row, 6, 'Delta-Wye')
        ws_tx.cell(row, 7, '5.75%')
        ws_tx.cell(row, 8, 'ONAN')
    
    for col in range(1, len(headers) + 1):
        ws_tx.column_dimensions[get_column_letter(col)].width = 18
    
    # === SHEET 6: LV SWITCHBOARDS (optional) ===
    if data.get('lv_swbd'):
        ws_lv = wb.create_sheet("LV Switchboards")
        headers = ["Tag", "Type", "Voltage", "Bus Rating", "Main Breaker", "Number of Feeders", "Feeds"]
        row = 1
        for col, header in enumerate(headers, start=1):
            cell = ws_lv.cell(row, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for swbd in data.get('lv_swbd', []):
            row += 1
            ws_lv.cell(row, 1, swbd.get('id', 'LV-SWBD'))
            ws_lv.cell(row, 2, swbd.get('type', 'Low Voltage Switchboard'))
            ws_lv.cell(row, 3, swbd.get('voltage', '480V'))
            ws_lv.cell(row, 4, swbd.get('bus_rating', '3200A'))
            ws_lv.cell(row, 5, swbd.get('main_breaker', '3200A'))
            ws_lv.cell(row, 6, len(swbd.get('feeders', [])))
            ws_lv.cell(row, 7, ', '.join(swbd.get('feeders', [])))
        
        for col in range(1, len(headers) + 1):
            ws_lv.column_dimensions[get_column_letter(col)].width = 20
    
    # === SHEET 7: UPS SYSTEMS ===
    ws_ups = wb.create_sheet("UPS Systems")
    
    headers = ["Tag", "Type", "Rating", "Input Voltage", "Output Voltage", "Battery Type", "Runtime", "Topology"]
    row = 1
    for col, header in enumerate(headers, start=1):
        cell = ws_ups.cell(row, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for ups in data.get('ups', []):
        row += 1
        ws_ups.cell(row, 1, ups['id'])
        ws_ups.cell(row, 2, ups.get('type', 'STATIC'))
        ws_ups.cell(row, 3, f"{ups.get('kw', 'N/A')}kW / {ups.get('kva', 'N/A')}kVA")
        ws_ups.cell(row, 4, '480V')
        ws_ups.cell(row, 5, '480V')
        ws_ups.cell(row, 6, ups.get('battery', 'LI-ION'))
        ws_ups.cell(row, 7, '15 min at full load')
        ws_ups.cell(row, 8, 'Double Conversion')
    
    for col in range(1, len(headers) + 1):
        ws_ups.column_dimensions[get_column_letter(col)].width = 18
    
    # === SHEET 8: BILL OF MATERIALS ===
    ws_bom = wb.create_sheet("Bill of Materials")
    
    headers = ["Item", "Description", "Manufacturer", "Model/Part Number", "Quantity", "Unit", "Unit Price", "Total Price"]
    row = 1
    for col, header in enumerate(headers, start=1):
        cell = ws_bom.cell(row, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    bom_items = [
        (1, "Diesel Generator", "Cummins", "C4000 D5", len(data.get('generators', [])), "EA", "TBD", "TBD"),
        (2, "13.8kV Switchgear", "ABB", "UniGear ZS1", len(data.get('mv_switchboards', [])), "EA", "TBD", "TBD"),
        (3, "Ring Main Unit", "Schneider Electric", "SM6-36", len(data.get('rmu', [])), "EA", "TBD", "TBD"),
        (4, "Dry Type Transformer 200kVA", "General Electric", "9T21B3874", len(data.get('transformers', [])), "EA", "TBD", "TBD"),
        (5, "480V Switchboard", "Eaton", "POW-R-LINE C", len(data.get('lv_swbd', [])), "EA", "TBD", "TBD"),
        (6, "Modular UPS 1MW", "Schneider Electric", "Galaxy VX", len(data.get('ups', [])), "EA", "TBD", "TBD"),
        (7, "Medium Voltage Cable (15kV)", "Southwire", "MV-105", 1000, "FT", "TBD", "TBD"),
        (8, "Low Voltage Cable (600V)", "Southwire", "THHN/THWN", 5000, "FT", "TBD", "TBD"),
        (9, "Cable Terminations (MV)", "3M", "Cold Shrink", 32, "EA", "TBD", "TBD"),
        (10, "Grounding System", "Various", "Ground Rods, Wire", 1, "LOT", "TBD", "TBD"),
    ]
    
    for item_data in bom_items:
        row += 1
        for col, value in enumerate(item_data, start=1):
            ws_bom.cell(row, col, value)
    
    ws_bom.column_dimensions['B'].width = 30
    ws_bom.column_dimensions['C'].width = 20
    ws_bom.column_dimensions['D'].width = 20
    
    # Save workbook
    wb.save(output_file)
    
    print(f"✅ Equipment Schedule created: {output_file}")
    print(f"   Sheets: {', '.join(wb.sheetnames)}")
    return output_file


if __name__ == "__main__":
    # Test with existing metadata
    test_metadata = "pachyderm_metadata.json"
    if Path(test_metadata).exists():
        print("Generating equipment schedule...")
        output = create_equipment_schedule(test_metadata, "saga_pryor_equipment_schedule.xlsx")
        print(f"\n✅ Equipment schedule generated successfully!")
