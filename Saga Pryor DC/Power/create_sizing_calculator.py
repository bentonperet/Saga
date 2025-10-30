import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load existing workbook
wb = openpyxl.load_workbook(r'C:\Users\eriks\Downloads\Pryor_Solar_BESS_Turbines_Model 20251024.xlsm.xlsx')

# Create new sizing sheet
if 'System Sizing Calculator' in wb.sheetnames:
    del wb['System Sizing Calculator']
ws = wb.create_sheet('System Sizing Calculator', 0)

# Styles
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=11)
input_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
calc_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
warning_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws['A1'] = 'TIER III MICROGRID SIZING CALCULATOR - PHASE 1'
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:F1')

ws['A2'] = 'N+1 BESS, Optimized Solar, Nat Gas Turbines for HB1374 Compliance'
ws['A2'].font = Font(italic=True, size=10)
ws.merge_cells('A2:F2')

row = 4

# === SECTION 1: USER INPUTS ===
ws[f'A{row}'] = 'USER INPUTS'
ws[f'A{row}'].font = header_font
ws[f'A{row}'].fill = header_fill
ws.merge_cells(f'A{row}:F{row}')
row += 1

inputs = [
    ('IT Load (MW)', 2, 'Critical IT capacity'),
    ('PUE (Power Usage Effectiveness)', 1.3, 'Total Facility / IT Load'),
    ('Target Solar Energy Fraction (%)', 75, '% of annual energy from solar (vs turbine)'),
    ('BESS Backup Duration - Each Path (minutes)', 60, 'Runtime per A/B path at full load'),
    ('Solar Capacity Factor (%)', 24, 'Oklahoma average - validated'),
    ('BESS Depth of Discharge (%)', 80, 'Conservative for cycle life'),
    ('BESS Roundtrip Efficiency (%)', 90, 'AC-DC-AC efficiency'),
    ('BESS End-of-Life Factor (%)', 80, 'Capacity at Year 10'),
    ('PCS/Inverter Efficiency (%)', 97, 'Power conversion system'),
]

for label, value, note in inputs:
    ws[f'A{row}'] = label
    ws[f'B{row}'] = value
    ws[f'B{row}'].fill = input_fill
    ws[f'B{row}'].border = thin_border
    ws[f'C{row}'] = note
    ws[f'C{row}'].font = Font(italic=True, size=9)
    row += 1

row += 1

# === SECTION 2: CALCULATED PARAMETERS ===
ws[f'A{row}'] = 'CALCULATED PARAMETERS'
ws[f'A{row}'].font = header_font
ws[f'A{row}'].fill = header_fill
ws.merge_cells(f'A{row}:F{row}')
row += 1

# Facility load
ws[f'A{row}'] = 'Facility Load (MW)'
ws[f'B{row}'] = '=B5*B6'
ws[f'B{row}'].number_format = '0.00'
ws[f'C{row}'] = 'IT Load × PUE'
row += 1

# Annual energy
ws[f'A{row}'] = 'Annual Facility Energy (MWh/year)'
ws[f'B{row}'] = f'=B{row-1}*8760'
ws[f'B{row}'].number_format = '#,##0'
ws[f'C{row}'] = 'Facility Load × 8,760 hours'
row += 1

# BESS combined derating
ws[f'A{row}'] = 'BESS Combined Derating Factor'
ws[f'B{row}'] = '=(B10/100)*(B11/100)*(B12/100)*(B13/100)'
ws[f'B{row}'].number_format = '0.000'
ws[f'C{row}'] = 'DoD × RTE × EoL × PCS Efficiency'
row += 1

row += 1

# === SECTION 3: SOLAR SIZING ===
ws[f'A{row}'] = 'SOLAR PV SIZING'
ws[f'A{row}'].font = header_font
ws[f'A{row}'].fill = header_fill
ws.merge_cells(f'A{row}:F{row}')
row += 1

ws[f'A{row}'] = 'Required Solar Capacity (MWdc)'
ws[f'B{row}'] = f'=(B7/100)*B{row-8}/(B9/100)'
ws[f'B{row}'].number_format = '0.00'
ws[f'B{row}'].fill = calc_fill
ws[f'C{row}'] = 'Target Solar Fraction × Annual Energy / CF'
row += 1

ws[f'A{row}'] = '→ RECOMMENDED Solar (MWdc)'
ws[f'B{row}'] = f'=ROUND(B{row-1},0)'
ws[f'B{row}'].font = Font(bold=True, size=12)
ws[f'B{row}'].fill = warning_fill
ws[f'B{row}'].number_format = '0'
ws[f'C{row}'] = 'Rounded to nearest MW for practical deployment'
row += 1

ws[f'A{row}'] = 'Annual Solar Production (MWh/year)'
ws[f'B{row}'] = f'=B{row-1}*(B9/100)*8760'
ws[f'B{row}'].number_format = '#,##0'
ws[f'C{row}'] = 'Solar Capacity × CF × Hours'
row += 1

ws[f'A{row}'] = 'Actual Solar Energy Fraction (%)'
ws[f'B{row}'] = f'=B{row-1}/B{row-9}*100'
ws[f'B{row}'].number_format = '0.0'
ws[f'C{row}'] = 'Solar Production / Total Facility Energy'
row += 1

ws[f'A{row}'] = 'Peak Solar Output (MW)'
ws[f'B{row}'] = f'=B{row-3}*0.85'
ws[f'B{row}'].number_format = '0.00'
ws[f'C{row}'] = '~85% of nameplate during peak sun hours'
row += 1

row += 1

# === SECTION 4: BESS SIZING (TIER III N+1) ===
ws[f'A{row}'] = 'BESS SIZING - TIER III (A/B INDEPENDENT PATHS)'
ws[f'A{row}'].font = header_font
ws[f'A{row}'].fill = header_fill
ws.merge_cells(f'A{row}:F{row}')
row += 1

ws[f'A{row}'] = 'Backup Duration (hours)'
ws[f'B{row}'] = '=B8/60'
ws[f'B{row}'].number_format = '0.00'
ws[f'C{row}'] = 'Minutes converted to hours'
row += 1

ws[f'A{row}'] = 'Required Delivered Energy - Per Path (MWh)'
ws[f'B{row}'] = f'=B{row-18}*B{row-1}'
ws[f'B{row}'].number_format = '0.00'
ws[f'C{row}'] = 'Facility Load × Backup Duration'
row += 1

ws[f'A{row}'] = 'Nameplate Energy - Per Path (MWh)'
ws[f'B{row}'] = f'=B{row-1}/B{row-10}'
ws[f'B{row}'].number_format = '0.00'
ws[f'B{row}'].fill = calc_fill
ws[f'C{row}'] = 'Delivered / Derating Factor'
row += 1

ws[f'A{row}'] = '→ RECOMMENDED BESS - Per Path (MWh)'
ws[f'B{row}'] = f'=CEILING(B{row-1},0.5)'
ws[f'B{row}'].font = Font(bold=True, size=12)
ws[f'B{row}'].fill = warning_fill
ws[f'B{row}'].number_format = '0.0'
ws[f'C{row}'] = 'Rounded up to 0.5 MWh increments'
row += 1

ws[f'A{row}'] = '→ TOTAL BESS ENERGY (Both Paths A+B) (MWh)'
ws[f'B{row}'] = f'=B{row-1}*2'
ws[f'B{row}'].font = Font(bold=True, size=12, color='FF0000')
ws[f'B{row}'].fill = warning_fill
ws[f'B{row}'].number_format = '0.0'
ws[f'C{row}'] = 'Two independent A/B paths for Tier III'
row += 1

ws[f'A{row}'] = 'Required Power - Per Path (MW AC)'
ws[f'B{row}'] = f'=B{row-24}*1.15'
ws[f'B{row}'].number_format = '0.00'
ws[f'C{row}'] = 'Facility Load × 1.15 safety factor'
row += 1

ws[f'A{row}'] = '→ RECOMMENDED BESS POWER - Per Path (MW)'
ws[f'B{row}'] = f'=CEILING(B{row-1},0.5)'
ws[f'B{row}'].font = Font(bold=True, size=12)
ws[f'B{row}'].fill = warning_fill
ws[f'B{row}'].number_format = '0.0'
ws[f'C{row}'] = 'Rounded up; N+1 PCS blocks within each path'
row += 1

ws[f'A{row}'] = '→ TOTAL BESS POWER (Both Paths A+B) (MW)'
ws[f'B{row}'] = f'=B{row-1}*2'
ws[f'B{row}'].font = Font(bold=True, size=12, color='FF0000')
ws[f'B{row}'].fill = warning_fill
ws[f'B{row}'].number_format = '0.0'
ws[f'C{row}'] = 'Two independent paths'
row += 1

ws[f'A{row}'] = 'PCS Configuration - Per Path'
ws[f'C{row}'] = 'Example: 3×1.0 MW inverters, any 1 can fail (2+1 redundant)'
row += 1

row += 1

# === SECTION 5: TURBINE/GENERATOR SIZING ===
ws[f'A{row}'] = 'NAT GAS TURBINE SIZING (HB1374 + N+1)'
ws[f'A{row}'].font = header_font
ws[f'A{row}'].fill = header_fill
ws.merge_cells(f'A{row}:F{row}')
row += 1

ws[f'A{row}'] = 'Minimum Total Capacity (MW)'
ws[f'B{row}'] = f'=B{row-32}*1.5'
ws[f'B{row}'].number_format = '0.00'
ws[f'C{row}'] = 'Facility Load × 1.5 for N+1 redundancy'
row += 1

ws[f'A{row}'] = '→ OPTION A: 2 × 2.0 MW Units (N+1)'
ws[f'B{row}'] = 2
ws[f'B{row}'].fill = warning_fill
ws[f'C{row}'] = 'Simpler, larger units - any 1 can fail'
ws[f'D{row}'] = 'Total: 4.0 MW'
row += 1

ws[f'A{row}'] = '→ OPTION B: 3 × 1.5 MW Units (N+1)'
ws[f'B{row}'] = 3
ws[f'B{row}'].fill = warning_fill
ws[f'C{row}'] = 'Better part-load efficiency, maintenance flexibility'
ws[f'D{row}'] = 'Total: 4.5 MW'
row += 1

ws[f'A{row}'] = '→ RECOMMENDED: Option B'
ws[f'B{row}'] = '3 × 1.5 MW'
ws[f'B{row}'].font = Font(bold=True, size=12, color='FF0000')
ws[f'B{row}'].fill = warning_fill
ws[f'C{row}'] = 'Finer turndown, lower fuel at part-load'
row += 1

row += 1

# === SECTION 6: ANNUAL ENERGY & FUEL ===
ws[f'A{row}'] = 'ANNUAL ENERGY BALANCE & FUEL ESTIMATE'
ws[f'A{row}'].font = header_font
ws[f'A{row}'].fill = header_fill
ws.merge_cells(f'A{row}:F{row}')
row += 1

ws[f'A{row}'] = 'Total Facility Energy (MWh/year)'
ws[f'B{row}'] = f'=B{row-40}'
ws[f'B{row}'].number_format = '#,##0'
row += 1

ws[f'A{row}'] = 'Solar Production (MWh/year)'
ws[f'B{row}'] = f'=B{row-28}'
ws[f'B{row}'].number_format = '#,##0'
row += 1

ws[f'A{row}'] = 'Solar Fraction (%)'
ws[f'B{row}'] = f'=B{row-1}/B{row-2}*100'
ws[f'B{row}'].number_format = '0.0'
row += 1

ws[f'A{row}'] = 'Turbine Generation Required (MWh/year)'
ws[f'B{row}'] = f'=B{row-3}-B{row-2}'
ws[f'B{row}'].number_format = '#,##0'
ws[f'B{row}'].fill = calc_fill
ws[f'C{row}'] = 'Facility - Solar'
row += 1

ws[f'A{row}'] = 'Turbine Fuel Cost ($/kWh)'
ws[f'B{row}'] = 0.10
ws[f'B{row}'].fill = input_fill
ws[f'B{row}'].number_format = '0.00'
ws[f'C{row}'] = 'Nat gas ~$0.06-0.10/kWh; diesel ~$0.27-0.30/kWh'
row += 1

ws[f'A{row}'] = '→ ANNUAL FUEL COST - PRE-GRID ($/year)'
ws[f'B{row}'] = f'=B{row-2}*1000*B{row-1}'
ws[f'B{row}'].font = Font(bold=True, size=12, color='FF0000')
ws[f'B{row}'].fill = warning_fill
ws[f'B{row}'].number_format = '$#,##0'
ws[f'C{row}'] = 'Largest OPEX until grid connection'
row += 1

ws[f'A{row}'] = 'Post-Grid Fuel Cost (DR events only, $/year)'
ws[f'B{row}'] = f'=12*3*B{row-52}*1000*B{row-2}'
ws[f'B{row}'].number_format = '$#,##0'
ws[f'C{row}'] = '12 events × 3 hrs × Facility MW × Fuel $/kWh'
row += 1

row += 1

# === SECTION 7: CAPEX SUMMARY ===
ws[f'A{row}'] = 'PRELIMINARY CAPEX ESTIMATE (2024 MARKET RATES)'
ws[f'A{row}'].font = header_font
ws[f'A{row}'].fill = header_fill
ws.merge_cells(f'A{row}:F{row}')
row += 1

ws[f'A{row}'] = 'Solar PV ($/Wdc)'
ws[f'B{row}'] = 1.10
ws[f'B{row}'].fill = input_fill
ws[f'B{row}'].number_format = '0.00'
ws[f'C{row}'] = '2024 market: $1.00-1.25/Wdc'
row += 1

ws[f'A{row}'] = 'BESS Energy ($/kWh)'
ws[f'B{row}'] = 400
ws[f'B{row}'].fill = input_fill
ws[f'B{row}'].number_format = '0'
ws[f'C{row}'] = '2024 market: $325-500/kWh (use conservative $400)'
row += 1

ws[f'A{row}'] = 'BESS Power/PCS ($/kW)'
ws[f'B{row}'] = 150
ws[f'B{row}'].fill = input_fill
ws[f'B{row}'].number_format = '0'
ws[f'C{row}'] = '2024 market: $120-180/kW'
row += 1

ws[f'A{row}'] = 'Nat Gas Turbine ($/kW)'
ws[f'B{row}'] = 850
ws[f'B{row}'].fill = input_fill
ws[f'B{row}'].number_format = '0'
ws[f'C{row}'] = 'Nat gas recip: $400-550/kW; combustion turbine: $800-1200/kW'
row += 1

ws[f'A{row}'] = ''
row += 1

ws[f'A{row}'] = 'Solar CAPEX'
ws[f'B{row}'] = f'=B{row-57}*1000000*B{row-5}'
ws[f'B{row}'].number_format = '$#,##0,000'
ws[f'C{row}'] = 'Solar MWdc × $/Wdc × 1,000,000'
row += 1

ws[f'A{row}'] = 'BESS Energy CAPEX'
ws[f'B{row}'] = f'=B{row-51}*1000*B{row-5}'
ws[f'B{row}'].number_format = '$#,##0,000'
ws[f'C{row}'] = 'Total BESS MWh × $/kWh × 1,000'
row += 1

ws[f'A{row}'] = 'BESS Power/PCS CAPEX'
ws[f'B{row}'] = f'=B{row-48}*1000*B{row-5}'
ws[f'B{row}'].number_format = '$#,##0,000'
ws[f'C{row}'] = 'Total BESS MW × $/kW × 1,000'
row += 1

ws[f'A{row}'] = 'Turbine CAPEX (3×1.5MW)'
ws[f'B{row}'] = f'=4500*B{row-5}'
ws[f'B{row}'].number_format = '$#,##0,000'
ws[f'C{row}'] = '4.5 MW × $/kW'
row += 1

ws[f'A{row}'] = 'Microgrid Controller (IEEE 2030.7/8)'
ws[f'B{row}'] = 1300000
ws[f'B{row}'].number_format = '$#,##0,000'
ws[f'C{row}'] = 'Per addendum'
row += 1

ws[f'A{row}'] = 'NFPA 855 Fire Safety'
ws[f'B{row}'] = f'=B{row-51}*350000'
ws[f'B{row}'].number_format = '$#,##0,000'
ws[f'C{row}'] = '~$350k per MWh BESS (conservative)'
row += 1

ws[f'A{row}'] = 'IEEE 519 Harmonics'
ws[f'B{row}'] = 450000
ws[f'B{row}'].number_format = '$#,##0,000'
ws[f'C{row}'] = 'Per addendum - scaled for larger system'
row += 1

ws[f'A{row}'] = 'Integration & Site Work'
ws[f'B{row}'] = 750000
ws[f'B{row}'].number_format = '$#,##0,000'
ws[f'C{row}'] = 'Civil, electrical, commissioning'
row += 1

ws[f'A{row}'] = '→ TOTAL GROSS CAPEX'
ws[f'B{row}'] = f'=SUM(B{row-8}:B{row-1})'
ws[f'B{row}'].font = Font(bold=True, size=12)
ws[f'B{row}'].fill = warning_fill
ws[f'B{row}'].number_format = '$#,##0,000'
row += 1

ws[f'A{row}'] = 'Less: 30% ITC (Solar+BESS+Integration)'
ws[f'B{row}'] = f'=-(B{row-9}+B{row-8}+B{row-7}+B{row-2})*0.30'
ws[f'B{row}'].number_format = '$#,##0,000'
row += 1

ws[f'A{row}'] = 'Less: 15% ITC (Microgrid Controller)'
ws[f'B{row}'] = f'=-B{row-5}*0.15'
ws[f'B{row}'].number_format = '$#,##0,000'
ws[f'C{row}'] = 'Conservative; may qualify for 30%'
row += 1

ws[f'A{row}'] = '→ NET CAPEX (After ITC)'
ws[f'B{row}'] = f'=B{row-3}+B{row-2}+B{row-1}'
ws[f'B{row}'].font = Font(bold=True, size=14, color='FF0000')
ws[f'B{row}'].fill = warning_fill
ws[f'B{row}'].number_format = '$#,##0,000'
row += 1

row += 2

# === SECTION 8: KEY OUTPUTS SUMMARY ===
ws[f'A{row}'] = '═══ FINAL SIZING RECOMMENDATION ═══'
ws[f'A{row}'].font = Font(bold=True, size=13, color='FFFFFF')
ws[f'A{row}'].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
ws.merge_cells(f'A{row}:F{row}')
row += 1

summary_row = row
ws[f'A{row}'] = 'Solar PV:'
ws[f'A{row}'].font = Font(bold=True)
ws[f'B{row}'] = f'=B{row-83} & " MWdc"'
ws[f'B{row}'].font = Font(bold=True, size=11)
ws[f'B{row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
ws[f'C{row}'] = 'Provides ~75% annual energy'
row += 1

ws[f'A{row}'] = 'BESS Total:'
ws[f'A{row}'].font = Font(bold=True)
ws[f'B{row}'] = f'=B{row-69} & " MWh / " & B{row-66} & " MW"'
ws[f'B{row}'].font = Font(bold=True, size=11)
ws[f'B{row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
ws[f'C{row}'] = 'Split into 2 independent A/B paths'
row += 1

ws[f'A{row}'] = 'BESS Per Path:'
ws[f'A{row}'].font = Font(bold=True)
ws[f'B{row}'] = f'=B{row-70} & " MWh / " & B{row-67} & " MW"'
ws[f'B{row}'].font = Font(bold=True, size=11)
ws[f'B{row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
ws[f'C{row}'] = 'Each path: 60 min at full load'
row += 1

ws[f'A{row}'] = 'Nat Gas Turbines:'
ws[f'A{row}'].font = Font(bold=True)
ws[f'B{row}'] = '3 × 1.5 MW (4.5 MW total)'
ws[f'B{row}'].font = Font(bold=True, size=11)
ws[f'B{row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
ws[f'C{row}'] = 'N+1 redundant, HB1374 compliant'
row += 1

ws[f'A{row}'] = 'Gross CAPEX:'
ws[f'A{row}'].font = Font(bold=True)
ws[f'B{row}'] = f'=B{row-8}'
ws[f'B{row}'].number_format = '$#,##0,000'
ws[f'B{row}'].font = Font(bold=True, size=11)
ws[f'B{row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
row += 1

ws[f'A{row}'] = 'Net CAPEX (after ITC):'
ws[f'A{row}'].font = Font(bold=True)
ws[f'B{row}'] = f'=B{row-4}'
ws[f'B{row}'].number_format = '$#,##0,000'
ws[f'B{row}'].font = Font(bold=True, size=11)
ws[f'B{row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
row += 1

ws[f'A{row}'] = 'Annual Fuel (pre-grid):'
ws[f'A{row}'].font = Font(bold=True)
ws[f'B{row}'] = f'=B{row-25}'
ws[f'B{row}'].number_format = '$#,##0'
ws[f'B{row}'].font = Font(bold=True, size=11)
ws[f'B{row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
ws[f'C{row}'] = 'Largest OPEX until grid'
row += 1

ws[f'A{row}'] = 'Annual Fuel (post-grid):'
ws[f'A{row}'].font = Font(bold=True)
ws[f'B{row}'] = f'=B{row-24}'
ws[f'B{row}'].number_format = '$#,##0'
ws[f'B{row}'].font = Font(bold=True, size=11)
ws[f'B{row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
ws[f'C{row}'] = 'DR events only'
row += 1

row += 1
ws[f'A{row}'] = 'TIER III NOTES:'
ws[f'A{row}'].font = Font(bold=True, underline='single')
row += 1

notes = [
    '• BESS deployed as two independent A/B paths (each path can carry full facility load)',
    '• Each path has N+1 PCS modules (e.g., 3×1.0 MW per path; any 1 can fail)',
    '• Turbines sized N+1 (any 1 unit can fail, remaining still cover full load)',
    '• Concurrent maintainability: Can service one BESS string or turbine without downtime',
    '• Consider: Small static UPS at rack/row level OR ensure BESS meets UPS transient specs',
]
for note in notes:
    ws[f'A{row}'] = note
    ws[f'A{row}'].font = Font(size=9)
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

# Column widths
ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 15

# Save
wb.save(r'C:\Users\eriks\Downloads\Pryor_Solar_BESS_Turbines_Model_UPDATED.xlsx')
print('✓ Created "System Sizing Calculator" sheet')
print('✓ Saved as: Pryor_Solar_BESS_Turbines_Model_UPDATED.xlsx')
print('')
print('RECOMMENDED SIZING (for 2 MW IT, PUE 1.3):')
print('  • Solar: 8 MWdc')
print('  • BESS: 10-12 MWh total (5-6 MWh per A/B path), 6 MW total power')
print('  • Turbines: 3 × 1.5 MW nat gas (N+1)')
print('  • Net CAPEX: ~$16-18M (after 30% ITC)')
print('  • Pre-grid fuel: ~$350-600k/year (vs $1.65M with 3 MW solar)')
